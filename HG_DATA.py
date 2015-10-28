#!/sur/bin/env  python2
#-*-coding=utf-8-*-

'''
Description     :  海关数据更新xlsx 、rec文件生成程序
require :openpyxl
author shizhongxian@126.com
usage  $python HG_DATA.py  -f  haiguan_2015_08_25.txt
haiguan_2015_08_25.txt 文件为从kbase导出的rec文件
'''


import os
import string
import operator
import  sys
import shutil
import errno
from openpyxl import Workbook
from openpyxl import load_workbook
from optparse import OptionParser
#import  Draw_Cells_Line as DCL
import Rec2Table

type = sys.getfilesystemencoding()


def load_dict(tdfile,key_col,value_col_list):
    """
              根据文件和列来构造dict数据结构
    tdfile: 纯文本 表格样式的文件,列之间用"\t"分割 
    key_col:key列号,从0开始
    value_col_lilst:充当value的列号，列表形式[1,2,4],列号必须递增 
    """
    result = {}
    fin = open(tdfile)
    line_no = 0
    line = fin.readline()
    line = line.strip()
    if len(line.split("\t"))-1 < value_col_list[-1] or len(line.split("\t"))-1 < key_col :
        print "输入的列号大于文件列号".decode("utf-8").encode(type)
        sys.exit() 
    while line:
        line_no += 1
        if line_no%500==0:
            print "加载数据 ".decode("utf-8").encode(type),line_no," "
        items = line.split("\t")
        if len(items)-1 < value_col_list[-1] or len(items)-1 < key_col:
            print line," 列数小于输入的列数".decode("utf-8").encode(type)
        else:
            if result.has_key(items[key_col]):
                pass
            else:
                tmp_list = []
                for i in value_col_list:
                    tmp_list.append(items[i])
                result[items[key_col]] = tmp_list
        line = fin.readline()
    fin.close()
    return result


def mkdir_p(path):
    try:
        os.makedirs(path)
    except OSError as exc: # Python >2.5
        if exc.errno == errno.EEXIST and os.path.isdir(path):
            pass
        else: raise
        
def unit_trans(src_unit,tar_unit,num):
    '''
    
    '''
    unit_dict={'百万美元||千美元':1000.0,'千美元||百万美元':0.001,"吨||万吨":0.0001,"万吨||顿":10000,"万台||台":10000,"台||万台":0.0001}
    tmp_s = src_unit+"||"+tar_unit
    if unit_dict.has_key(tmp_s):
        return float(num)*unit_dict[tmp_s]
    else:
        print src_unit,tar_unit,"单位转化失败".decode("utf-8").encode(type)
        return False
    

def copyanything(src, dst):
    try:
        shutil.copytree(src, dst)
    except OSError as exc: # python >2.5
        if exc.errno == errno.ENOTDIR:
            shutil.copy(src, dst)
        else: raise
        
def copy_and_overwrite(from_path, to_path):
    if os.path.exists(to_path):
        shutil.rmtree(to_path)
    shutil.copytree(from_path, to_path)

def get_year_and_month(startyear,startmonth,endyear,endmonth):
    '''
    
    '''
    year_month_list = []
    for i in range(startyear,endyear):
        for j in range(startmonth,13):
            year_month_list.append(str(i)+"年-"+str(j)+"月")
    for j in range(1,endmonth+1):
            year_month_list.append(str(endyear)+"年-"+str(j)+"月")
    return year_month_list

def get_file_from_dir(dirname):
    '''
    返回目录下的文件名
    相对路径
    '''
    file_list = []
    for root,dirs,filenames in os.walk(dirname):
        for filename in filenames:
            file_list.append(os.path.join(root,filename))
    return file_list

def replace_growth(s):
    '''
    
    '''
    return string.replace(s, "同比增长", '')


def replace_im_ex(s):
    #HG20 
    if s.find("进口") != -1:
        tmp_str =  string.replace(s, "进口","进出口")
        #HG23
        if tmp_str.find("境内目的地") != -1:
            return string.replace(tmp_str, "境内目的地","境内目的地/货源地总额")
        else:
            return tmp_str
    #Hg20 
    elif s.find("出口") != -1:
        tmp_str =   string.replace(s, "出口","进出口")
    #HG23 
        if tmp_str.find("境内货源地") != -1:
            return  string.replace(tmp_str, "境内货源地","境内目的地/货源地总额")
        else:
            return tmp_str
    else:
        return s

def sorteddict(d):
    '''
    
    '''
    return [(k,d[k]) for k in sorted(d.keys())]


def get_rules(rulefile_name):
    '''
    
    '''
    rules_dict = {}
    with open(rulefile_name) as f:
        lines = f.readlines()[1:]
        for line in lines:
            line = line.strip()
            item_list = line.split(':')
            if len(item_list)==2:
                k = item_list[0]
                rule = item_list[1]
                if rules_dict.has_key(k):
                    print "规则中有重复,请检查".decode("utf-8").encode(type),rulefile_name
                else:
                    rules_dict.setdefault(k,rule)
    return rules_dict

def read_data(fname):
    '''
    
    '''
    data_dict = {}
    with open(fname) as f:
        lines = f.readlines()
        for line in lines:
            item_list = line.split("\t")
            if len(item_list) ==  8:
                indicator,area,area_code,year,month,num,unit=item_list[:-1]
                tmp_list = [indicator,area_code,year,month]
                key = "||".join(tmp_list)
                if data_dict.has_key(key):
                    data_dict[key] = [num,unit]
                else:
                    data_dict.setdefault(key,[num,unit])
    return data_dict
                    

def save_table_data(table_file_name,indicator_dir):
    '''
    
    '''
    file_list = get_file_from_dir(indicator_dir)
    
    #fin = open("2010_2015_allmonth_table.txt")
    fin = open(table_file_name)
    data = fin.readlines()[1:]
    
    for filename in file_list:
        print "正在保存数据:".decode('utf-8').encode(type),filename
        #filename = "HG_INDICATOR/HG1.txt"
        out_file_name = "HG_CLS_DATA/"+ os.path.basename(filename)[:-4]+"_data.txt"
        #out_file_name = "HG_CLS_DATA/"+ "test"+"_data.txt"
        fout = open(out_file_name,"w")
        with open(filename) as f:
            lines = f.readlines()[1:]
            line_no = 0
            for line in lines:
                line_no += 1
                if line_no%10 == 0:
                    sys.stdout.write("抽取指标数:".decode('utf-8').encode(type)+str(line_no)+"\r")
                    #print line_no,"\r"
                line = line.strip()
                slist = line.split("=")
                if len(slist) == 2:
                    indicator = slist[1]
                    for item in data:
                        #line_no += 1
                        #if line_no % 100000==0:
                            #print line_no
                        item_list = item.split("\t")
                        #print indicator,"----"
                        #print item_list[0],"----"
                        if indicator == item_list[0]:    
                            item = item.strip()
                            fout.write(item+"\t"+"stats"+"\n")
        fout.close()
    fin.close()
    
def explor_growth_indicator(indicator_dir):
    '''
    '''
    no_growth_indicators_tables = []
    file_list = get_file_from_dir(indicator_dir)
    for filename in file_list:
        #filename ="HG_INDICATOR/HG1.txt"
        print "processing:",filename
        with open(filename) as f:
            lines = f.readlines()[1:]
            base_indicators = set()
            growth_indicator = set()
            for line in lines:
                line = line.strip()
                item_list = line.split("=")
                if len(item_list) == 2:
                    indicator = item_list[1]
                    if indicator.find("同比增长") == -1:
                        base_indicators.add(indicator)
                    else:
                        growth_indicator.add(indicator)
        #替换‘同比增长’字符串
        growth_indicator = map(replace_growth, growth_indicator)
        growth_indicator = set(growth_indicator)
        if len(base_indicators) == len(growth_indicator):
            print "同比指标和基本指标个数相等".decode("utf-8").encode(type),len(growth_indicator)
            if base_indicators.issubset(growth_indicator) and growth_indicator.issubset(base_indicators):
                print "同比指标和基本指标完全相同".decode("utf-8").encode(type)
            else:
                print "同比指标和基本指标不完全相同".decode("utf-8").encode(type)
        else:
            print "同比指标个数不等于基本指标个数".decode("utf-8").encode(type)
            print "同比指标个数:".decode("utf-8").encode(type),len(growth_indicator)
            print "基本指标个数:".decode("utf-8").encode(type),len(base_indicators)
            no_growth_indicators_tables.append(filename)
    print '\n'.join(no_growth_indicators_tables)
    
    
def generate_up_value(indicator_dir,start_year,start_month,end_year,end_month):
    '''
        对没有同比数据的指标产生同比数据
    '''
    file_list = get_file_from_dir(indicator_dir)
    #file_list = ['HG_CLS_DATA/HG1_data.txt']
    
    for filename in file_list:
        print filename
        #对HG20_data  HG23_data文件,计算分地区进出口综合
        if filename.find('HG20_data') !=-1 or filename.find('HG23_data') !=-1:
            #计算地区进出口数据
            location_trade(filename,start_year,start_month,end_year,end_month)
#         #if filename.find('HG23_data') !=-1:
#         #    location_trade(filename)
            
        #以指标 地区 地区代码 月份为Key
        #以年份 数值 为value建立字典
        base_data_dict = {}
        indicator_set = set()
        with open(filename) as f:
            lines = f.readlines()
            for line in lines:
                line = line.strip()
                item_list = line.split('\t')
                #判断行内格式是否正确
                if len(item_list) == 8:
                    indicator = item_list[0].strip('[S]')
                    indicator_set.add(indicator)
                    area,area_code,year,month,num,unit = item_list[1:-1]
                    key = indicator+'@'+area+'@'+area_code+'@'+month
                    value_tmp_list = [year,num,unit]
                    #以指标 地区  地区代码 月为key ,年份 数值 单位为value建立词典 求同比数据
                    if base_data_dict.has_key(key):
                        base_data_dict[key].append(value_tmp_list)
                    else:
                        base_data_dict.setdefault(key,[])
        #找到需要计算同比增长的指标
        #indicator_list = base_data_dict.keys()
        base_indicators = set()
        growth_indicators = set()
        for indicator in indicator_set:
            if indicator.find("同比增长") == -1:
                #基本指标
                base_indicators.add(indicator)
            else:
                #同比增长指标
                growth_indicators.add(indicator)
        #calculate_indicators = set()
        growth_indicators = map(replace_growth, growth_indicators)
        growth_indicators = set(growth_indicators)
        #需要计算同比增长的指标集
        calculate_indicators = set()
        if len(base_indicators) == len(growth_indicators):
            #print "同比指标和基本指标个数相等",len(growth_indicators)
            if base_indicators.issubset(growth_indicators) and growth_indicators.issubset(base_indicators):
                #print "同比指标和基本指标完全相同"
                pass
            else:
                #print "同比指标和基本指标不完全相同"
                calculate_indicators = base_indicators - growth_indicators
                #print calculate_indicators
        else:
            #print "同比指标个数不等于基本指标个数"
            #print "同比指标个数:",len(growth_indicators)
            #print "基本指标个数:",len(base_indicators)
            calculate_indicators = base_indicators - growth_indicators
                #print '\n'.join(calculate_indicators)
            #按照月份来计算同比增长
        cal_resut = []
        if len(calculate_indicators) > 0:
            for key,value_list in base_data_dict.iteritems():
                tmp_indicator = key.split('@')[0]
                if tmp_indicator in calculate_indicators:
                    num_dict = {}
                    #单月数据 对应的年份、数值、单位
                    for year_item in value_list:
                        if len(year_item) == 3:
                            year,num,unit = year_item
                            year_no = year.strip('年')
                            #不同年对应的数值和单位词典
                            num_dict[string.atoi(year_no)] = [string.atof(num),unit]
                        #对词典 按照年份排序 生成序列
                    sorted_list = sorteddict(num_dict)
                    up_nums = []
                    for i in range(len(sorted_list)-1):
                        #判断年份是否相差为1年
                        if sorted_list[i+1][0]-sorted_list[i][0]==1:
                            cur_cal_year = sorted_list[i+1][0]
                            #list
                            cur_num_unit = sorted_list[i+1][1]
                            #list
                            last_num_unit = sorted_list[i][1]
                                #判断单位是否相同
                            if cur_num_unit[1] == last_num_unit[1]:
                                ratio_num = round((cur_num_unit[0]/last_num_unit[0]-1)*100,2)
                                #print (cur_num_unit[0]/last_num_unit[0]-1)*100
                                
                                up_nums.append([cur_cal_year,ratio_num])
                            else:
                                print key,cur_cal_year,"单位不统一"
                        #保存计算得到的数据
                    out_indicator,out_area,out_area_code,out_month = key.split('@')
                    out_indicator += "同比增长[S]"
                    for item in up_nums:
                        out_year = item[0]
                        out_num = item[1]
                        cal_resut.append([out_indicator,out_area,out_area_code,str(out_year)+"年",out_month,str(out_num),"%","calculated"])
            #outfilename = "/home/jay/workspace_new/HG_DATA/HG_CAL_DATA/" + os.path.basename(filename)[:-4]+"_calculated.txt"
            outfilename = filename
            #fout = open(outfilename,'w')
            #输出
            fout = open(outfilename,'a+')        
            for line in cal_resut:
                fout.write('\t'.join(line)+"\n")
            fout.close()    
        #
        print filename,"calculate end!"
        
def trade_top(filename ,start_year,start_month,end_year,end_month):
    '''
        计算月度顺差与逆差地区数据，并排序
    '''
    f = open(filename)
    lines = f.readlines()
    outfile_name = "HG_CLS_DATA/HG8_data.txt"
    fout = open(outfile_name,"w")
    outfile_name_uf = "HG_CLS_DATA/HG9_data.txt"
    fout_uf = open(outfile_name_uf,"w")
    year_months = get_year_and_month(start_year,start_month,end_year,end_month)
    for tmp_time in year_months:
        item_list = tmp_time.split('-')
        if len(item_list) == 2:
            cal_year = item_list[0]
            cal_month = item_list[1]
            sys.stdout.write(cal_year+cal_month+"\r")
            year_num = string.atoi(cal_year.strip("年"))
            last_year = str(year_num-1)+"年"
            #filename = 'HG_CLS_DATA/HG7_data.txt'
            #国家 [出口 ,进口]
            cur_data_dict = {}
            last_data_dict = {}
            #indicator_set = set()
            with open(filename) as f:
                lines = f.readlines()
            for line in lines:
                line = line.strip()
                item_list = line.split('\t')
                #判断行内格式是否正确
                if len(item_list) == 8:
                    area,area_code,year,month,num,unit = item_list[1:-1]
                    indicator = item_list[0]
                    #过滤掉组织  洲   联盟等地区
                    if year == cal_year and month == cal_month and indicator.find('组织')==-1 and indicator.find('联盟')==-1  and indicator.find('洲')==-1:
                        indicator = item_list[0]
                        pos1 = indicator.find(',')
                        pos2 = indicator.find(')[')
                        #获取地区
                        trade_area = indicator[pos1+1:pos2]
                        #获取进出口值
                        if cur_data_dict.has_key(trade_area):
                            if indicator.find('出口') != -1:
                                cur_data_dict[trade_area][0] = string.atof(num)
                            elif indicator.find('进口') != -1:
                                cur_data_dict[trade_area][1] = string.atof(num)
                        else:
                            cur_data_dict.setdefault(trade_area,['-','-'])
                            if indicator.find('出口') != -1:
                                cur_data_dict[trade_area][0] = string.atof(num)
                            elif indicator.find('进口') != -1:
                                cur_data_dict[trade_area][1] = string.atof(num)
                    #获取去年当期进出口值
                    elif year==last_year and month == cal_month:
                        indicator = item_list[0]
                        pos1 = indicator.find(',')
                        pos2 = indicator.find(')[')
                        trade_area = indicator[pos1+1:pos2]
                        if last_data_dict.has_key(trade_area):
                            if indicator.find('出口') != -1:
                                last_data_dict[trade_area][0] = string.atof(num)
                            elif indicator.find('进口') != -1:
                                last_data_dict[trade_area][1] = string.atof(num)
                        else:
                            last_data_dict.setdefault(trade_area,['-','-'])
                            if indicator.find('出口') != -1:
                                last_data_dict[trade_area][0] = string.atof(num)
                            elif indicator.find('进口') != -1:
                                last_data_dict[trade_area][1] = string.atof(num)
            #顺差
            cur_tmp_list = []
            for k,v in cur_data_dict.iteritems():
                if len(v) == 2 and v[0]!='-' and v[1]!='-':
                    cur_tmp_list.append((k,v[0]-v[1]))
            #去年当期
            last_tmp_dict = {}
            for k,v in last_data_dict.iteritems():
                if len(v) == 2 and v[0]!='-' and v[1]!='-':
                    last_tmp_dict[k] = v[0]-v[1]
                
            #outfile_name = "HG_CLS_DATA/" + cal_year + "_" + cal_month + "_" + "trade_area_data.txt"
            #fout = open("HG_CAL_DATA/trade_area_data.txt","w")
            #fout = open(outfile_name,"w")
            #顺差从小到大排序
            cur_tmp_list.sort(key=operator.itemgetter(1))
            #顺差国家  
            fb_area = cur_tmp_list[-10:]
            i = 0
            for item in fb_area:
                i +=1 
                fout.write("国别"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+item[0]+"\t"+"$"+"\tcalculated\n")
                fout.write("顺差额"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+str(item[1])+"\t"+"千美元"+"\tcalculated\n")
                fout.write("上年同期顺差额"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+str(last_tmp_dict[item[0]])+"\t"+"千美元"+"\tcalculated\n")
                #fout.write(item[0]+"\t"+str(item[1])+"\t"+str(last_tmp_dict[item[0]])+"\n")
            #逆差国家
            ub_area =cur_tmp_list[:10] 
            i = 0
            for item in ub_area:
                i +=1
                if last_tmp_dict.has_key(item[0]):
                    fout_uf.write("国别"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+item[0]+"\t"+"$"+"\tcalculated\n")
                    fout_uf.write("逆差额"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+str(abs(item[1]))+"\t"+"千美元"+"\tcalculated\n")
                    fout_uf.write("上年同期逆差额"+str(i)+"\t"+area+"\t" + area_code + "\t"+cal_year+"\t"+cal_month+"\t"+str(abs(last_tmp_dict[item[0]]))+"\t"+"千美元"+"\tcalculated\n")
                #tmp_list.sort(key=operator.itemgetter(1))
    fout_uf.close()
    fout.close()
    f.close()
    print "顺差 逆差国别数据计算结束 ! "
        
    
def location_trade(filename,start_year,start_month,end_year,end_month):
    '''
        分所在地，根据进口、出口数据计算进出口数据
    '''
    #print filename
    #sys.stdout.write("计算商品出口总额经营单位所在地数据".decode('utf-8').encode(type))
    sys.stdout.write("计算商品出口总额经营单位所在地数据 \n")
    year_months = get_year_and_month(start_year,start_month,end_year,end_month)
    for tmp_time in year_months:
        item_list = tmp_time.split('-')
        if len(item_list) == 2:
            cal_year = item_list[0]
            cal_month = item_list[1]
            sys.stdout.write(cal_year+cal_month+"\r")
            cal_data = []
            with open(filename) as f:
                line = f.readline()
                while line:
                    #过滤掉同比数据
                    if line.find("同比增长") == -1:
                        item_list = line.split("\t")
                        if len(item_list) == 8:
                            indicator,area,area_code,year,month,num,unit = item_list[0:-1]
                            #获取特定年份、月份数据
                            if year == cal_year and month == cal_month:
                                #HG20
                                if indicator == "商品出口总额(经营单位所在地)[S]" or indicator == "商品进口总额(经营单位所在地)[S]":
                                    #为指标添加标记，方便计算不同地区的 '商品出口总额(经营单位所在地)[S]' 计算完成后取消
                                    indicator = area + '@' +  indicator
                                else:
                                    pass
                                if indicator == "商品进口总额(境内目的地)[S]" or indicator == "商品出口总额(境内货源地)[S]":
                                    #为指标添加标记，方便计算不同地区的 '商品出口总额(经营单位所在地)[S]' 计算完成后取消
                                    indicator = area + '@' +  indicator
                                else:
                                    pass    
                                #保存需要计算的指标和数值
                                cal_data.append([indicator,area,area_code,year,month,num,unit])
                    line = f.readline()
            ex_im_num_dict = {}
            for tmp_list in cal_data:
                if len(tmp_list) == 7:
                    indicator,area,area_code,year,month,num,unit = tmp_list
                    #进口 出口 合并为一个指标
                    ex_im_indicator = replace_im_ex(indicator)
                    #初始化对应的出口 进口值
                    ex_im_num_dict.setdefault(ex_im_indicator,['-','-',area,area_code,year,month,num,unit])
                    #当前地区添加出口数据
                    if indicator.find("出口") !=-1:
                        ex_im_num_dict[ex_im_indicator][0] = string.atof(num)
                    #当前地区添加进口数据
                    elif indicator.find("进口") !=-1:
                        ex_im_num_dict[ex_im_indicator][1] = string.atof(num)
                    else:
                        print indicator,"指标不包含进口 出口"
            #输出数据         
            fout = open(filename,"a+")
            #fout = open("test.txt","a+")   
            for k,value_list in ex_im_num_dict.iteritems():
                ex_num,im_num,area,area_code,year,month,num,unit = value_list
                if ex_num !='-' and im_num !='-':
                    ex_im_sum = ex_num + im_num
                    out_list = [area,area_code,year,month,str(ex_im_sum),unit,"caculated"]
                    #去掉标记
                    pos = k.find('@')
                    if pos != -1:
                        k = k[pos+1:]
                    fout.write(k + "\t" + '\t'.join(out_list)+"\n")
            fout.close()
    sys.stdout.write("所在地进出口数据计算结束 ! ")
    
        

def data_to_excel(cal_year,cal_month,cal_data_file_name,cal_rule_name,xls_name):
    '''
    
    '''
    rules = get_rules(cal_rule_name)
    data = read_data(cal_data_file_name)
    #tmp_i = 0
    #for k_t,v_t in data.iteritems():
                #tmp_i += 1
                #if tmp_i==10:
                    #break
                #print k_t
    #for k_t,v_t in data.iteritems():
        #print k_t,v_t
    wb = load_workbook(xls_name)
    sheetnames = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetnames[0])
    #按照年月修改表格的标题
    exlTitle=ws.cell('A1').value
    #print exlTitle.__class__
    dataCh=cal_year + cal_month
    #dataCh=dataCh.decode('utf-8')    
    dataIni='xxxx年xx月'  
    #dataIni=dataIni.decode('utf-8')
    #转化年度数据标题格式
    if cal_data_file_name.find('HG10') != -1 or cal_data_file_name.find('HG11') != -1 or cal_data_file_name.find('HG12') != -1 or cal_data_file_name.find('HG13') != -1:
        dataCh = cal_year
        dataIni = 'xxxx年'
    dataCh=dataCh.decode('utf-8')    
    dataIni=dataIni.decode('utf-8')
    ws.cell('A1').value=exlTitle.replace(dataIni,dataCh)
    
    for cellname,rule in rules.iteritems():
        rule_item_list = rule.split("||")
        if len(rule_item_list) == 6:
            #变换规则
            s_indicator,s_areacode,s_year,s_month,s_unit,s_flag = rule_item_list
            if s_year == "default":
                s_year = cal_year
            if s_month == "default":
                s_month = cal_month
            #累计月份转化
            if s_flag == "L":
                if s_month== "1月":
                    pass
                else:
                    s_month = "1-"+s_month
            s_tmp_list = [s_indicator,s_areacode,s_year,s_month]
            s_rule = "||".join(s_tmp_list) 
            #if s_rule.find("同比增长") != -1:
                #print s_rule
            #print s_rule

            if data.has_key(s_rule):
                num,unit = data[s_rule]
                if s_unit.strip() == unit.strip():
                    ws.cell(cellname).value = num 
                else:
                    trans_num = unit_trans(unit, s_unit, num)
                    if trans_num:
                        ws.cell(cellname).value = trans_num
                    else:
                        print s_rule,"请查看单位转化问题",xls_name
                    #print s_unit,unit,"需要转换单位"
            else:
                ws.cell(cellname).value = '-'
    #DCL.Draw_Cells_Line(ws)
    wb.save(xls_name)
    
    
def save_to_xls(start_year,start_month,end_year,end_month):
    '''
    
    '''
    year_months = get_year_and_month(start_year,start_month,end_year,end_month)
    #逐月处理
    for tmp_time in year_months:
        item_list = tmp_time.split('-')
        if len(item_list) == 2:
            cur_year = item_list[0]
            cur_month = item_list[1]

            print tmp_time.decode("utf-8").encode(type)
            #cur_save_dir_name = "XLS_DATA/"+cur_year.strip("年")+"/"+ cur_month.strip("月") + "/"
            cur_save_dir_name = "XLS_DATA"+ os.path.sep + cur_year.strip("年")+os.path.sep + cur_month.strip("月") 

            basenames = []
            try:
                #复制文件到当月文件夹
                #copyanything('XLS_MODULE/',cur_save_dir_name)
                copy_and_overwrite('XLS_MODULE/', cur_save_dir_name)
            except Exception,e:
                print "文件夹复制失败",e
                sys.exit()
            else:
                print cur_save_dir_name,"文件夹复制成功 ".decode("utf-8").encode(type)
                xls_file_list = os.listdir(cur_save_dir_name)
                #获取xlsx文件名的基本名,切去掉后缀名
                for xls_name in xls_file_list:
                    basenames.append(os.path.basename(xls_name).strip(".xlsx"))
            #basenames = ['HG1','HG2']
            for bname in basenames:
                data_file_name = os.path.join("HG_CLS_DATA/",bname+'_data.txt')
                rule_file_name = os.path.join("RULES/",bname+'.txt')
                xls_file_name   = os.path.join(cur_save_dir_name,bname+'.xlsx')
                data_to_excel(cur_year, cur_month, data_file_name, rule_file_name,xls_file_name)
            print tmp_time.decode("utf-8").encode(type),"转化完毕".decode("utf-8").encode(type)
            
def to_rec():
    '''
    
    '''
    fout = open("HG_HTML_REC.txt","w")
    files_list = get_file_from_dir("HG_XLSX_0629/")
    data_dict = load_dict("FILECODE.TXT_table.txt", 0, [1])
    for filename in files_list:
        bname = os.path.basename(filename)
        name = bname.strip(".htm")
        #截取年份和月份
        pos1 = filename.find("201")
        pos2 = filename.rfind("HG")
        year_month = filename[pos1:pos2-1]
        year,month = year_month.split('/')
        if len(month) == 1:
            month = "0"+month
        year_month = year+month
        if data_dict.has_key(name):
            REC_FILENAME = data_dict[name][0]+"_"+ year_month
            REC_FILECODE = name+"_" + year_month
            REC_HTMLTABLECODE = ""
            with open(filename) as f:
                REC_HTMLTABLECODE = f.read()
            #REC_HTMLTABLECODE = REC_HTMLTABLECODE.decode('utf-8')    
            fout.write("<REC>")
            fout.write("<FILENAME>="+REC_FILENAME+"\n")
            fout.write("<FILECODE>="  +  REC_FILECODE+"\n")
            fout.write("<HTMLTABLECODE>="+REC_HTMLTABLECODE+"\n")
    fout.close()
                
                
def generate_Rec(cal_year='2015',cal_month='5'):
    '''
    
    '''
    long_month = ''
    if len(cal_month) == 1:
        long_month = '0'+cal_month
    else:
        long_month = cal_month
        
    cjfyhg_htmlpreview = "Rec/CJFYHG_HTMLPREVIEW.txt"
    fin = open(cjfyhg_htmlpreview)
    out_filename = cjfyhg_htmlpreview.strip(".txt")+cal_year+"_" + cal_month + ".rec"
    fout = open(out_filename,"w")
    lines = fin.readlines()
    for line in lines:
        if line.find( "ym") != -1:
            fout.write(string.replace(line, "ym", cal_year +long_month))
            continue
        else:
            fout.write(line)
    fout.close()
    fin.close()
    
    main_monthly_hot_name = "Rec/CJFYHG_MAIN_MONTHLY_HOT.txt"
    fin = open(main_monthly_hot_name)
    out_filename = main_monthly_hot_name.strip(".txt") + cal_year +"_" + cal_month + ".rec"
    fout = open(out_filename,"w")
    lines = fin.readlines()
    for line in lines:
        if line.find( "ym") != -1:
            fout.write(string.replace(line, "ym", cal_year +long_month))
            continue
        elif line.find("year") !=-1:
            fout.write(string.replace(line, "year", cal_year))
            continue
        elif line.find("month") != -1:
            fout.write(string.replace(line, "month", cal_month))
            continue
        else:
            fout.write(line)
    fout.close()
    fin.close()
    
    
    sub_monthly_hot_name = "Rec/CJFYHG_SUBJECT_MONTHLY_HOT.txt"
    fin = open(sub_monthly_hot_name)
    out_filename = sub_monthly_hot_name.strip(".txt") + cal_year +"_" + cal_month + ".rec"
    fout = open(out_filename,"w")
    lines = fin.readlines()
    for line in lines:
        #fout.write(string.replace(line, "xxxx_x", cal_year +"_"+cal_month))
        if line.find("year") != -1:
            fout.write(string.replace(line, "year", cal_year))
            continue
        elif line.find("month") != -1:
            fout.write(string.replace(line, "month", long_month))
            continue
        else:
            fout.write(line)
    fout.close()
    fin.close()
    
if __name__ == "__main__":
    optparser = OptionParser()
    optparser.add_option('-f', '--inputFile',
                         dest='input',
                         help='filename of haiguan  rec data',
                         default=None)
    (options, args) = optparser.parse_args()
    inFile = None
    if options.input is None:
            inFile = sys.stdin
            #inFile = "INTEGRATED-DATASET.csv"
    elif options.input is not None:
            inFile = options.input
    else:
            print 'No dataset filename specified, system with exit\n'
            sys.exit('System will exit')
    
    start_year,start_month,end_year,end_month = [2011,1,2015,6]
    tb_name = Rec2Table.Rec2Table(inFile, 'hg_table.txt')
    save_table_data(tb_name,"HG_INDICATOR/")
    generate_up_value("HG_CLS_DATA",start_year,start_month,end_year,end_month)
    trade_top( 'HG_CLS_DATA/HG7_data.txt',start_year,start_month,end_year,end_month)
    save_to_xls(start_year,start_month,end_year,end_month)
    generate_Rec(cal_year=str(end_year),cal_month=str(end_month))
    print "End!"
    
    
    
    
    
